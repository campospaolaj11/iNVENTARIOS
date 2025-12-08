import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

def crear_excel_desde_json(json_file="data_dashboard.json", excel_file="inventario_completo.xlsx"):
    """
    Crear Excel profesional desde datos JSON (simulando los datos del dashboard)
    """
    
    # Datos de ejemplo (puedes reemplazar con lectura de JSON)
    productos_data = [
        {'codigo': 'PROD001', 'nombre': 'Tornillo M8x20', 'categoria': 'Ferretería', 'stock_actual': 150, 'stock_minimo': 50, 'costo_unitario': 0.15, 'precio_venta': 0.30, 'ubicacion': 'A-01'},
        {'codigo': 'PROD002', 'nombre': 'Tuerca M8', 'categoria': 'Ferretería', 'stock_actual': 200, 'stock_minimo': 50, 'costo_unitario': 0.10, 'precio_venta': 0.20, 'ubicacion': 'A-02'},
        {'codigo': 'PROD003', 'nombre': 'Arandela M8', 'categoria': 'Ferretería', 'stock_actual': 30, 'stock_minimo': 50, 'costo_unitario': 0.05, 'precio_venta': 0.10, 'ubicacion': 'A-03'},
        {'codigo': 'PROD004', 'nombre': 'Cable 2x14 AWG', 'categoria': 'Eléctricos', 'stock_actual': 500, 'stock_minimo': 100, 'costo_unitario': 1.50, 'precio_venta': 3.00, 'ubicacion': 'B-01'},
        {'codigo': 'PROD005', 'nombre': 'Interruptor Simple', 'categoria': 'Eléctricos', 'stock_actual': 80, 'stock_minimo': 20, 'costo_unitario': 2.50, 'precio_venta': 5.00, 'ubicacion': 'B-02'},
    ]
    
    # Crear DataFrame
    df = pd.DataFrame(productos_data)
    
    # Calcular columnas adicionales
    df['valor_stock'] = df['stock_actual'] * df['costo_unitario']
    df['estado'] = df.apply(lambda row: 'CRÍTICO' if row['stock_actual'] <= row['stock_minimo'] 
                                       else 'BAJO' if row['stock_actual'] <= row['stock_minimo'] * 1.5 
                                       else 'NORMAL', axis=1)
    df['pedido_sugerido'] = df.apply(lambda row: max(0, row['stock_minimo'] - row['stock_actual']), axis=1)
    
    # Crear Excel con formato profesional
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        # Hoja 1: Inventario completo
        df.to_excel(writer, sheet_name='Inventario', index=False)
        
        # Hoja 2: Productos críticos
        df_criticos = df[df['estado'] == 'CRÍTICO']
        df_criticos.to_excel(writer, sheet_name='Stock Crítico', index=False)
        
        # Hoja 3: Resumen por categoría
        df_resumen = df.groupby('categoria').agg({
            'stock_actual': 'sum',
            'stock_minimo': 'sum',
            'valor_stock': 'sum'
        }).reset_index()
        df_resumen.columns = ['Categoría', 'Stock Total', 'Stock Mínimo', 'Valor Total']
        df_resumen.to_excel(writer, sheet_name='Resumen por Categoría', index=False)
        
        # Hoja 4: KPIs
        kpis_data = {
            'Métrica': [
                'Total Productos',
                'Productos Críticos',
                'Valor Total Inventario',
                'Productos por Reabastecer'
            ],
            'Valor': [
                len(df),
                len(df_criticos),
                f"${df['valor_stock'].sum():,.2f}",
                len(df[df['pedido_sugerido'] > 0])
            ]
        }
        df_kpis = pd.DataFrame(kpis_data)
        df_kpis.to_excel(writer, sheet_name='KPIs', index=False)
    
    # Aplicar formato
    aplicar_formato_excel(excel_file)
    print(f"✅ Excel creado: {excel_file}")
    return excel_file

def aplicar_formato_excel(excel_file):
    """Aplicar formato profesional al Excel"""
    wb = load_workbook(excel_file)
    
    # Colores corporativos
    header_fill = PatternFill(start_color="0284c7", end_color="0284c7", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Formato del encabezado
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Bordes y alineación para todas las celdas
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')
        
        # Colorear estados en hoja de inventario
        if sheet_name == 'Inventario':
            estado_col = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == 'estado':
                    estado_col = idx
                    break
            
            if estado_col:
                for row in range(2, ws.max_row + 1):
                    estado_cell = ws.cell(row=row, column=estado_col)
                    if estado_cell.value == 'CRÍTICO':
                        estado_cell.fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
                        estado_cell.font = Font(color="dc2626", bold=True)
                    elif estado_cell.value == 'BAJO':
                        estado_cell.fill = PatternFill(start_color="fed7aa", end_color="fed7aa", fill_type="solid")
                        estado_cell.font = Font(color="ea580c", bold=True)
                    else:
                        estado_cell.fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
                        estado_cell.font = Font(color="059669", bold=True)
    
    wb.save(excel_file)

def importar_excel_a_json(excel_file="template_inventario.xlsx", json_file="productos_importados.json"):
    """
    Leer Excel y convertir a JSON para el dashboard
    """
    try:
        df = pd.read_excel(excel_file, sheet_name='Inventario')
        
        # Convertir a formato JSON compatible con el dashboard
        productos = []
        for _, row in df.iterrows():
            producto = {
                'codigo': str(row['codigo']),
                'nombre': str(row['nombre']),
                'categoria': str(row['categoria']),
                'stock_actual': int(row['stock_actual']),
                'stock_minimo': int(row['stock_minimo']),
                'costo_unitario': float(row['costo_unitario']),
                'precio_venta': float(row['precio_venta']),
                'ubicacion_bodega': str(row['ubicacion'])
            }
            productos.append(producto)
        
        # Guardar JSON
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(productos, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Importados {len(productos)} productos")
        print(f"✅ JSON guardado: {json_file}")
        return productos
        
    except Exception as e:
        print(f"❌ Error al importar: {e}")
        return None

def exportar_kpis_dashboard(json_file="../frontend/public/mock-data.json", excel_file="kpis_dashboard.xlsx"):
    """
    Exportar KPIs del dashboard a Excel para análisis
    """
    try:
        # Leer JSON del dashboard
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Crear hojas con diferentes análisis
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # KPIs generales
            kpis = data['kpis']
            df_kpis = pd.DataFrame([kpis])
            df_kpis.to_excel(writer, sheet_name='KPIs', index=False)
            
            # Productos
            df_productos = pd.DataFrame(data['productos'])
            df_productos.to_excel(writer, sheet_name='Productos', index=False)
            
            # Stock por categoría
            df_categorias = pd.DataFrame(data['stock_por_categoria'])
            df_categorias.to_excel(writer, sheet_name='Por Categoría', index=False)
            
            # Productos críticos
            df_criticos = pd.DataFrame(data['productos_criticos'])
            df_criticos.to_excel(writer, sheet_name='Críticos', index=False)
        
        aplicar_formato_excel(excel_file)
        print(f"✅ KPIs exportados: {excel_file}")
        return excel_file
        
    except Exception as e:
        print(f"❌ Error al exportar: {e}")
        return None

if __name__ == "__main__":
    print("🚀 Automatización Excel - Sistema de Inventarios\n")
    
    print("1️⃣  Creando Excel profesional desde datos...")
    crear_excel_desde_json()
    
    print("\n2️⃣  Importando Excel a JSON...")
    importar_excel_a_json()
    
    print("\n3️⃣  Exportando KPIs del dashboard...")
    exportar_kpis_dashboard()
    
    print("\n✅ ¡Proceso completado!")
    print("\nArchivos generados:")
    print("  📄 inventario_completo.xlsx - Excel profesional con múltiples hojas")
    print("  📄 productos_importados.json - Datos en formato JSON")
    print("  📄 kpis_dashboard.xlsx - KPIs exportados del dashboard")
