import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime

def crear_template_inventario(archivo_salida="template_inventario.xlsx"):
    """Crear plantilla Excel para carga de inventarios"""
    
    # Datos de ejemplo
    data = {
        'codigo': ['PROD001', 'PROD002', 'PROD003'],
        'nombre': ['Producto Ejemplo 1', 'Producto Ejemplo 2', 'Producto Ejemplo 3'],
        'categoria': ['Categoría A', 'Categoría B', 'Categoría A'],
        'stock_actual': [100, 50, 75],
        'stock_minimo': [20, 10, 15],
        'costo_unitario': [10.50, 25.00, 15.75],
        'precio_venta': [15.00, 35.00, 22.50],
        'ubicacion': ['A-01', 'B-02', 'A-03']
    }
    
    df = pd.DataFrame(data)
    
    # Crear Excel
    df.to_excel(archivo_salida, sheet_name='Inventario', index=False)
    
    # Aplicar formato
    wb = load_workbook(archivo_salida)
    ws = wb['Inventario']
    
    # Estilo del encabezado
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(archivo_salida)
    print(f"✅ Template creado: {archivo_salida}")

def validar_excel_inventario(archivo_path):
    """Validar que el Excel tenga las columnas correctas"""
    columnas_requeridas = [
        'codigo', 'nombre', 'categoria', 'stock_actual', 
        'stock_minimo', 'costo_unitario', 'precio_venta', 'ubicacion'
    ]
    
    try:
        df = pd.read_excel(archivo_path)
        columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
        
        if columnas_faltantes:
            print(f"❌ Faltan columnas: {', '.join(columnas_faltantes)}")
            return False
        
        print("✅ Excel válido")
        return True
        
    except Exception as e:
        print(f"❌ Error al validar: {e}")
        return False

if __name__ == "__main__":
    # Crear template
    crear_template_inventario()
    
    # Validar template creado
    validar_excel_inventario("template_inventario.xlsx")
